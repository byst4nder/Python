import sys

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
'''
命令行接受数字 N，在一个 Excel 电子表格中创建一个 N×N 的乘法表。例如，如果这样执行程序： 
 py multiplicationTable.py 6 
 '''
def createMultip(num:int):
    # 创建excel文件
    wb = openpyxl.Workbook()
    sh = wb.active

    # 定义格式
    font = Font(bold=True)
    # 根据输入数字 创建乘法表
    for row in range(0,num+1):
        for column in range(0,num+1):
            if row+1 == 1:
                sh.cell(row=row+1,column=column+1).value = column
                # 给表头设置格式
                sh.cell(row=row+1, column=column+1).font = font
            elif column+1 == 1:
                sh.cell(row=row+1,column=column+1).value = row
                # 给表头设置格式
                sh.cell(row=row+1,column=column+1).font = font
            else:
                sh.cell(row=row+1,column=column+1).value = sh.cell(row=1,column=column+1).value * sh.cell(row=row+1,column=1).value

    for row in range(0, num + 1):
        for column in range(0, num + 1):
            print(str(sh.cell(row=row+1,column=column+1).value),end='      ')
        print()

    # 保存文件
    wb.save('multip.xlsx')



if __name__ == '__main__':

    if len(sys.argv) == 2:
        createMultip(int(sys.argv[1]))
    else:
        print("use: python multiplicationTable [num]")

